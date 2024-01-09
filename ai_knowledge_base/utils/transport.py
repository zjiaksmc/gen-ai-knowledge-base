import io
import openpyxl
from xlsx2html.core import *


def worksheet_to_data(ws, locale=None, fs=None, default_cell_border="none"):
    merged_cell_map = {}
    if OPENPYXL_24:
        merged_cell_ranges = ws.merged_cell_ranges
        excluded_cells = set(ws.merged_cells)
    else:
        merged_cell_ranges = [cell_range.coord for cell_range in ws.merged_cells.ranges]
        excluded_cells = set(
            [
                cell
                for cell_range in merged_cell_ranges
                for rows in rows_from_range(cell_range)
                for cell in rows
            ]
        )

    for cell_range in merged_cell_ranges:
        try:
            cell_range_list = list(ws[cell_range])
            m_cell = cell_range_list[0][0]
            colspan = len(cell_range_list[0])
            rowspan = len(cell_range_list)
            merged_cell_map[m_cell.coordinate] = {
                "attrs": {
                    "colspan": None if colspan <= 1 else colspan,
                    "rowspan": None if rowspan <= 1 else rowspan,
                },
                "cells": [c for rows in cell_range_list for c in rows],
            }
            try:
                excluded_cells.remove(m_cell.coordinate)
            except KeyError:
                pass
        except TypeError:
            pass
    
    max_col_number = 0
    data_list = []
    for row_i, row in enumerate(ws.iter_rows()):
        data_row = []
        data_list.append(data_row)
        for col_i, cell in enumerate(row):
            row_dim = ws.row_dimensions[cell.row]
            if cell.coordinate in excluded_cells or row_dim.hidden:
                continue
            if col_i > max_col_number:
                max_col_number = col_i
            height = 19
            if row_dim.customHeight:
                height = round(row_dim.height, 2)
            f_cell = None
            if fs:
                f_cell = fs[cell.coordinate]
            try:
                cell_data = {
                    "column": cell.column,
                    "row": cell.row,
                    "value": cell.value,
                    "formatted_value": format_cell(cell, locale=locale, f_cell=f_cell),
                    # "attrs": {"id": get_cell_id(cell)},
                    "attrs": {},
                    "style": {"height": f"{height}pt"},
                }
            except:
                cell_data = {
                    "column": cell.column,
                    "row": cell.row,
                    "value": cell.value,
                    "formatted_value": cell.value,
                    # "attrs": {"id": get_cell_id(cell)},
                    "attrs": {},
                    "style": {"height": f"{height}pt"},
                }
            merged_cell_info = merged_cell_map.get(cell.coordinate, {})
            if merged_cell_info:
                cell_data["attrs"].update(merged_cell_info["attrs"])
            cell_data["style"].update(
                get_styles_from_cell(cell, merged_cell_info, default_cell_border)
            )
            data_row.append(cell_data)

    col_list = []
    max_col_number += 1

    column_dimensions = sorted(
        ws.column_dimensions.items(), key=lambda d: column_index_from_string(d[0])
    )

    for col_i, col_dim in column_dimensions:
        if not all([col_dim.min, col_dim.max]):
            continue
        width = 0.89
        if col_dim.customWidth:
            width = round(col_dim.width / 10.0, 2)
        col_width = 96 * width

        for _ in six.moves.range((col_dim.max - col_dim.min) + 1):
            max_col_number -= 1
            col_list.append(
                {
                    "index": col_dim.index,
                    "hidden": col_dim.hidden,
                    "style": {"width": "{}px".format(col_width)},
                }
            )
            if max_col_number < 0:
                break
    return {"rows": data_list, "cols": col_list, "images": images_to_data(ws)}


def render_table(data, append_headers, append_lineno, apply_style=False):
    html = [
        "<table>"
    ]
    hidden_columns = set()
    # html.append("<colgroup>")
    # for col in data["cols"]:
    #     if col["hidden"]:
    #         hidden_columns.add(col["index"])
    #     if apply_style:
    #         html.append(
    #             '<col {attrs} style="{styles}">'.format(
    #                 attrs=render_attrs(col.get("attrs")),
    #                 styles=render_inline_styles(col.get("style")),
    #             )
    #         )
    #     else:
    #         html.append(
    #             '<col {attrs}>'.format(
    #                 attrs=render_attrs(col.get("attrs"))
    #             )
    #         )
    # html.append("</colgroup>")

    append_headers(data, html)

    for i, row in enumerate(data["rows"]):
        trow = ["<tr>"]
        append_lineno(trow, i)
        for cell in row:
            if cell["column"] in hidden_columns:
                continue
            images = data["images"].get((cell["column"], cell["row"])) or []
            formatted_images = []
            for img in images:
                styles = render_inline_styles(img["style"])
                img_tag = (
                    '<img width="{width}" height="{height}"'
                    'style="{styles_str}"'
                    'src="{src}"'
                    "/>"
                ).format(styles_str=styles, **img)
                formatted_images.append(img_tag)
            if apply_style:
                trow.append(
                    (
                        '<td {attrs_str} style="{styles_str}">'
                        "{formatted_images}"
                        "{formatted_value}"
                        "</td>"
                    ).format(
                        attrs_str=render_attrs(cell["attrs"]),
                        styles_str=render_inline_styles(cell["style"]),
                        formatted_images="\n".join(formatted_images),
                        **cell,
                    )
                )
            else:
                trow.append(
                    (
                        '<td {attrs_str}>'
                        "{formatted_images}"
                        "{formatted_value}"
                        "</td>"
                    ).format(
                        attrs_str=render_attrs(cell["attrs"]),
                        formatted_images="".join(formatted_images),
                        **cell,
                    )
                )

        trow.append("</tr>")
        html.append("".join(trow))
    html.append("</table>")
    return "".join(html)


def xlsx2html(
    filepath,
    output=None,
    locale="en",
    parse_formula=False,
    append_headers=(lambda dumb1, dumb2: True),
    append_lineno=(lambda dumb1, dumb2: False),
    default_cell_border="none",
):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # get number of sheets
    if len(wb.worksheets)>0:
        html = ""
        for i, sheet in enumerate(wb.worksheets):
            fs = None
            if parse_formula:
                fb = openpyxl.load_workbook(filepath, data_only=False)
                fs = get_sheet(fb, i)
            ws = get_sheet(wb, i)
            data = worksheet_to_data(
                ws, locale=locale, fs=fs, default_cell_border=default_cell_border
            )
            html_data = render_table(data, append_headers, append_lineno)
            # combines multiple excel sheets into 1 file
            html = html + f"<h1>{sheet.title}</h1>\n" + html_data + "\n"
    else:
        fs = None
        if parse_formula:
            fb = openpyxl.load_workbook(filepath, data_only=False)
            fs = get_sheet(fb, 0)
        ws = get_sheet(wb, 0)
        data = worksheet_to_data(
            ws, locale=locale, fs=fs, default_cell_border=default_cell_border
        )
        html = render_table(data, append_headers, append_lineno)
    
    if not output:
        output = io.StringIO()
    if isinstance(output, str):
        output = open(output, "w")
    output.write(html)
    output.flush()
    return output

